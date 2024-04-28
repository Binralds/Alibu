from openpyxl import Workbook
from openpyxl import load_workbook
from Function import sakEkrn
import os





if os.path.exists("Dati.xlsx"):
    wb = load_workbook("Dati.xlsx")
    print("Atveru datni...")
else:
    ex = Workbook()
    ac = ex.active
    Top = [["ID", "NOSK.", "SKAITS"]]
    for row in Top:
        ac.append(row)
    print("Izveidoju jaunu datni...")
    ex.save("Dati.xlsx")



sakEkrn()
