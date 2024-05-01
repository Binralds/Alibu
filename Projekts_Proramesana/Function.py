
from openpyxl import load_workbook
import time







def sakEkrn():
    global ex
    global sht
    ex = load_workbook("Dati.xlsx")
    sht = ex["Sheet"]
    print("""Izvēlieties opciju : \n
    Produkti || Skaits || Rediģēt || Palīdzība || Iziet
    """)
    Izvele = str(input("Izvēle : " + "").lower())

    match str(Izvele):

        case "produkti":
            tirit()
            produkti()
        case "skaits":
            tirit()
            skaits()
        case "rediget":
            tirit()
            rediget()
        case "rediģēt":
            tirit()
            rediget()
        case "palīdzība":
            tirit()
            palidziba()
        case "palidziba":
            tirit()
            palidziba()
        case "iziet":
            iziet = str(input("Vai tiešām vēlaties iziet? Y vai N: " + " ")).lower()
            if iziet == "y":
                print("Pārtraucu darbību, visu labu!...")
                time.sleep(1.0)
                exit()
            elif iziet == "n":
                print("Atgriežu Jūs uz sākuma ekrānu...")
                time.sleep(1.5)
                sakEkrn()
            else:
                print("Nesaprotu atbildi, atgriežu Jūs uz sākuma ekrānu")
                time.sleep(1.5)
                sakEkrn()
        case _:
            print("Nederīga opcija, lūdzu mēģiniet vēlreiz...")
            time.sleep(1.0)
            sakEkrn()


def rediget():
    id = input("Ievadiet produkta ID :" + " ")

    if not id.isdigit():
        print("Nederīga vērtība, lūdzu ievadiet ID")
        time.sleep(1.0)
        sakEkrn()

    else:

        pass
    i = 0
    for row in sht:
        i = i + 1
        if str(ex.active.cell(row=i, column=1).value) == str(id):
            nosk = ex.active.cell(row=i, column=2).value
            skaits = ex.active.cell(row=i, column=3).value
            print("Tiks rediģets produkts " + str(nosk) + ", kura skaits ir " + str(skaits))
            atb = input("Turpināt? Y vai N : ").lower()



            if atb == "y":
                opc = input("Pievienot vai noņemt? :" + " ").lower()
                if opc == "pievienot":
                    piev = input("Cik daudz vēlaties pievienot?:" + " ")
                    if not piev.isdigit():
                        print("Nederīga vērtība, lūdzu ievadiet skaitu")
                        time.sleep(0.5)
                        sakEkrn()

                    else:
                        ex.active.cell(row=i, column=3).value = int(ex.active.cell(row=i, column=3).value) + int(piev)
                        print("Produkta " + str(nosk) + " jaunais daudzums ir " + str(ex.active.cell(row=i, column=3).value))
                        ex.save("Dati.xlsx")
                        velv = str(input("Vai vēlaties rediģēt vēlvienu produktu? Y vai N: " + " ")).lower()
                        if velv == "y":
                            tirit()
                            time.sleep(1.0)
                            rediget()
                        elif velv == "n":
                            print("Atgriežu Jūs uz sākuma ekrānu...")
                            time.sleep(1.5)
                            sakEkrn()
                        else:
                            print("Nesaprotu atbildi, atgriežu Jūs uz sākuma ekrānu")
                            time.sleep(1.5)
                            sakEkrn()
                elif opc == "nonemt" or "noņemt":
                    noNemt = input("Cik daudz vēlaties noņemt?:" + " ")
                    if not noNemt.isdigit():
                        print("Nederīga vērtība, lūdzu ievadiet SKAITU")
                        time.sleep(0.5)
                        sakEkrn()
                    else:
                        ex.active.cell(row=i, column=3).value = int(ex.active.cell(row=i, column=3).value) - int(noNemt)
                        time.sleep(0.5)
                        print("Produkta " + str(nosk) + " jaunais daudzums ir " + str(ex.active.cell(row=i, column=3).value))
                        ex.save("Dati.xlsx")
                        time.sleep(1.0)
                        velv = str(input("Vai vēlaties rediģēt vēlvienu produktu? Y vai N: " + " ")).lower()
                        if velv == "y":
                            tirit()
                            time.sleep(1.0)
                            rediget()
                        elif velv == "n":
                            print("Atgriežu Jūs uz sākuma ekrānu...")
                            time.sleep(1.5)
                            sakEkrn()
                        else:
                            print("Nesaprotu atbildi, atgriežu Jūs uz sākuma ekrānu")
                            time.sleep(1.5)
                            sakEkrn()


                else:
                    print("Nederīga vērtība")
                    sakEkrn()

            elif atb == "n":
                print("Novirzu atpakaļ uz sākuma ekrānu...")
                time.sleep(0.5)
                sakEkrn()

            else:
                print("Nesapratu, novirzu atpakaļ uz sākuma ekrānu...")
                time.sleep(1.0)
                sakEkrn()
    else:
        print("Nav atrasts produkts ar ID: " + id)
        time.sleep(1.0)
        sakEkrn()
    atb = input("Rediģēt citus produktus? Y vai N :" + " ").lower()
    if atb == "y":
        rediget()
    elif atb == "n":
        sakEkrn()
    else:
        print("Nesapratu, novirzu uz sākuma ekrānu...")
        time.sleep(0.5)
        sakEkrn()


def skaits():
    tirit()
    id = input("Ievadiet ID produktam, kura skaitu vēlaties apskatīt :" + " ")

    if not id.isdigit():
        print("Nederīga vērtība, lūdzu ievadiet derīgu ID")
        time.sleep(1.0)
        sakEkrn()


    else:

        pass
        i = 0
        for row in sht:
            i = i + 1
            if str(ex.active.cell(row=i, column=1).value) == str(id):
                nosk = sht.cell(row=i, column=2).value
                daudz = sht.cell(row=i, column=3).value
                print("Produkta " + str(nosk) + " daudzums ir " + str(daudz))
                break
        else:
            print("Ievadītais ID : " + str(id) + " nav atpzīts, lūdzu mēģiniet vēlreiz")
            time.sleep(0.5)
            sakEkrn()

    atb = input("Skatīt citus produktus? Y vai N:" + " ").lower()
    if atb == "y":
        skaits()
    elif atb == "n":
        sakEkrn()
    else:
        print("Nesapratu, novirzu uz sākuma ekrānu...")
        time.sleep(0.5)
        sakEkrn()


def produkti():

    print("""Izvēlieties darbību 
        Pievienot || Dzēst || Atrast || Atpakaļ
    """)
    izv = str(input("Jūsu izvēle:" + " ").lower())
    match izv:
        case "pievienot":
            prod_piev()
        case "dzēst":
            prod_dzest()
        case "dzest":
            prod_dzest()
        case "atrast":
            prod_atr()
        case "atpakal":
            tirit()
            sakEkrn()
        case "atpakaļ":
            tirit()
            sakEkrn()
        case _:
            print("Nesapratu, lūdzu mēģiniet vēlreiz...")
            produkti()


def prod_piev():
    tirit()
    max = sht.max_row + 1
    newID = input("Lūdzu ievadiet jaunā produkta ID:" + " ")
    if not newID.isdigit():
        print("Nederīga vērtība, lūdzu ievadiet vērtību, kas atbilst ID")
        time.sleep(0.5)
        produkti()
    nosk = input("Lūdzu ievadiet jaunā produkta NOSAUKUMU:" + " ")
    daudz = input("Lūdzu ievadiet jaunā produkta SKAITU:" + " ")
    if not daudz.isdigit():
        print("Nederīga vērtība, lūdzu ievadiet SKAITU")
        time.sleep(0.5)
        produkti()
    i = 1
    for row in range(sht.max_row-1):
        i = i + 1
        dupe = int(ex.active.cell(row=i, column=1).value)
        if str(dupe) == str(newID):
            print("Tads ID jau pastāv, duplicēšana radīs problēmas.")
            sakEkrn()
        else:
            pass


    print("Produktu ar ID : " + str(newID) + " ,nosaukumu: " + str(nosk) + " un daudzumu: " + str(
        daudz) + " pievienošu izklājlapai")
    time.sleep(0.5)
    atb = input("Turpināt? Y vai N :" + " ").lower()
    if atb == "n":
        print("Dzēšu datus...")
        time.sleep(0.3)
        print("Novirzu atpakaļ uz sākuma ekrānu")
        time.sleep(0.3)
        sakEkrn()
    elif atb == "y":
        sht.cell(row=max, column=1).value = newID
        sht.cell(row=max, column=2).value = nosk
        sht.cell(row=max, column=3).value = int(daudz)
        ex.save("Dati.xlsx")
        print("Izmaiņas saglabātas rindā " + str(max))
        time.sleep(0.5)
        izv = input("Vai vēlaties pievienot vēlvienu produktu? Y vai N " + "").lower()
        if izv == "y":
            prod_piev()
        elif izv == "n":
            print("Novirzu uz sākuma ekrānu...")
            sakEkrn()
        else:
            print("Nesapratu atbildi, novirzu uz sākuma ekrānu...")
            time.sleep(1.0)
            sakEkrn()
    else:
        print("Neizprotu atbildi, nosūtu uz sākuma ekrānu...")
        time.sleep(0.5)
        sakEkrn()


def prod_atr():
    tirit()

    aizm = str(input("Ievadiet nosaukumu produktam, kura ID/Skaitu vēlaties noskaidrot:" + " "))
    i = 0
    for row in sht:
        i = i+1
        if str(ex.active.cell(row=i, column=2).value) == aizm:
            id = str(ex.active.cell(row=i, column=1).value)
            skaits = str(ex.active.cell(row=i, column=3).value)
            print("Produkta" + " " + aizm + " " + "ID ir" + " " + id + " " + "un skaits ir" + " " + skaits)
            time.sleep(0.5)
            izv = input("Vai vēlaties atrast vēlvienu produktu? Y vai N: " + "").lower()
            if izv == "y":
                prod_piev()
            elif izv == "n":
                print("Novirzu uz sākuma ekrānu...")
                sakEkrn()
            else:
                print("Nesapratu atbildi, novirzu uz sākuma ekrānu...")
                time.sleep(1.0)
                sakEkrn()
    else:
        print("Nevarēju atrast produktu ar norādīto nosaukumu, lūdzu mēģiniet vēlreiz...")
        time.sleep(0.5)
        produkti()




def skatit():
    velv = input("Vai vēlaties apskatīt vēl kāda cita produkta skaitu? Y vai N" + " ")
    if velv == "y":
        tirit()
        skaits()
    elif velv == "n":
        print("Novadu uz sākuma ekrānu...")
        time.sleep(0.5)
        sakEkrn()
    else:
        print("Nesapratu atbildi, novirzu uz sākuma ekrānu...")
        time.sleep(0.5)
        sakEkrn()

def prod_dzest():
    tirit()
    id = input("Ievadiet tā produkta ID, kuru vēlaties dzēst:" + " ")
    if not id.isdigit():
        print("Nederīga vērtība, lūdzu ievadiet vērtību, kas atbilst ID")
        time.sleep(0.5)
        sakEkrn()
    i = 0
    for row in sht:
        i = i+1
        if str(ex.active.cell(row=i, column=1).value) == str(id):
            nosk = ex.active.cell(row=i, column=2).value
            daudzums = ex.active.cell(row=i, column=3).value
            print("Tiks dzēsts produkts " + str(nosk) + " ar daudzumu " + str(daudzums))
            atbilde = input("Vai tiešām vēlaties dzēst šo produktu? Datus atgūt nebūs iespējams. Y vai N: ").lower()
            if atbilde == "y":
                sht.delete_rows(i)
                print("Veiksmīgi izdzēsu rindu " + str(i))
                ex.save("Dati.xlsx")
                atbildeNext = input("Vai vēlaties dzēst vēl kādus datus? Y vai N: ").lower()
                if atbildeNext == "y":
                    prod_dzest()
                elif atbildeNext == "n":
                    print("Novirzu atpakaļ uz sākuma ekrānu...")
                    time.sleep(1.5)
                    sakEkrn()
                else:
                    print("Nesaprotu atbildi, novirzu uz sākuma ekrānu...")
                    time.sleep(1.5)
                    sakEkrn()
            elif atbilde == "n":
                print("Novadu uz sākuma ekrānu...")
                time.sleep(1.5)
                sakEkrn()
            else:
                print("Nesaprotu atbildi, novirzu uz sākuma ekrānu...")
                time.sleep(1.5)
                sakEkrn()
    else:
        print("Neatradu produktu ar doto ID: " + str(id) + " ,lūdzu mēģiniet vēlreiz")
        time.sleep(0.5)
        aiz = input("Vai aizmirsāt kāda produkta id?: Y  vai N " + " ").lower()
        if aiz == "y":
            print("Atļaujiet man jums palīdzēt, novadīšu Jūs uz produktu atrašanas funkciju.")
            time.sleep(0.5)
            prod_atr()
        elif aiz == "n":
            print("Novirzu uz sākuma ekrānu...")
            time.sleep(0.5)
            sakEkrn()
        else:
            print("Atbildi nesapratu, novirzu uz sākuma ekrānu...")
            time.sleep(0.5)
            sakEkrn()


def palidziba():
    print("""
    Ja Jums ir radušās problēmas ar kādas funkcijas izpildi, lūdzu apskatiet GITHUB izveidoto README failu ar nosaukumu "Palīdzība"
    
    Šajā failā esmu iekļāvis instrukciju, kā izmantot katru no sākumā redzamajām funkcijām,
    
    Kā arī kļūdu ziņojumu aprakstu. 
    """)

    atpakal = input("Kad esiet gatavi, ievadiet šeit jebko lai atgrieztos sākuma ekrānā: ")

    match atpakal:
        case _:
            print("Atgriežu Jūs uz sākuma ekrānu...")
            time.sleep(1.5)
            sakEkrn()

def tirit():
    i = 0
    while i < 10:
        i = i + 1
        print("""







        """)

# rahhhh
