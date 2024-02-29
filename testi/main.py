
import time
import timeit
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook

def tirit():
    i = 0
    while i < 10:
        i = i + 1
        print("""







        """)


def atrast():

    aizm = input("Ievadiet nosaukumu produktam, kura ID/Skaitu vēlaties noskaidrot:" + " ")

    print(time.time())
    i = 0
    for row in ac:
        i = i+1
        if ex.active.cell(row=i, column=2).value == aizm:
            id = str(ex.active.cell(row=i, column=1).value)
            skaits = str(ex.active.cell(row=i, column=3).value)
            print("Produkta" + " " + aizm + " " + "ID ir" + " " + id + " " + "un skaits ir" + " " + skaits)

    print(time.time())


#vidējais funkcijas izpildes laiks ir 0.007 sekundes



def sakEkrn():
    izvele=input("Ievadi kko:" + " ").lower()

    if izvele == "atrast":
        atrast()


if os.path.exists("Data.xlsx"):
    Book = load_workbook("Data.xlsx")
    print("Atver datni")
else:
    Book = Workbook()
    Top = [["ID", "NOSK.", "SKAITS"]]
    for row in Top:
        Book.active.append(row)
    Book.active.cell(row=1, column=999).value = 1
    print("Izveidota jauna datne")
    Book.save("Data.xlsx")



ex = load_workbook("Data.xlsx")
ac = ex["Sheet"]

sakEkrn()







